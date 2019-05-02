import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
wb.get_sheet_names()
# ['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet.title)

a=sheet['A1'].value
print(a)
c = sheet['B1']
print(c.value)
c.value
c.row
c.column


print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
# 'Row 1, Column B is Apples'
print( 'Cell ' + c.coordinate + ' is ' + c.value) 
# 'Cell B1 is Apples'
print(sheet['C1'].value)
# 73



>>> from openpyxl import Workbook
>>> wb = Workbook(write_only = True)
>>> ws = wb.create_sheet()
>>> from openpyxl.cell import WriteOnlyCell
>>> from openpyxl.comments import Comment
>>> from openpyxl.styles import Font
>>> cell = WriteOnlyCell(ws, value="hello world")
>>> cell.font = Font(name='Courier', size=36)
>>> cell.comment = Comment(text="A comment", author="Author's Name")
>>> ws.append([cell, 3.14, None])
>>> wb.save('write_only_file.xlsx')